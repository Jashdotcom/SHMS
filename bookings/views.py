from datetime import date, timedelta
import re

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.http import HttpResponse, HttpResponseForbidden
from django.urls import reverse
from django.utils.dateparse import parse_date
from django.utils import timezone
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from accounts.models import User
from hostel.models import Room
from core.emails import send_templated_email
from core.notifications import create_notification

from .forms import BookingForm
from .models import Booking


def _is_admin_user(user):
	return user.is_authenticated and (user.is_superuser or user.is_staff or user.role == User.ROLE_ADMIN)


def _get_export_bed_label(booking):
	room_number = (booking.room.room_number or "").strip()
	prefix = room_number.split("-")[0].strip() if room_number else ""
	prefix = prefix[:1].upper() if prefix else ""

	bed_number = (booking.bed.bed_number or "").strip()
	match = re.search(r"(\d+)", bed_number)
	bed_index = match.group(1) if match else "1"

	return f"{prefix}{bed_index}" if prefix else bed_number


def _get_latest_unique_bookings_for_export():
	# Keep only the newest record (highest id) for the same booking identity.
	bookings = Booking.objects.select_related("user", "room", "bed").order_by("-id")
	latest_by_key = {}

	for booking in bookings:
		key = (
			booking.user_id,
			booking.room_id,
			booking.bed_id,
			booking.start_date,
			booking.end_date,
		)
		if key not in latest_by_key:
			latest_by_key[key] = booking

	return list(latest_by_key.values())


@login_required
def book_room_view(request):
	if request.user.role != User.ROLE_STUDENT:
		return HttpResponseForbidden("Admins cannot book rooms")

	rooms = Room.objects.all().prefetch_related("beds").order_by("room_number")

	if request.method == "POST":
		form = BookingForm(request.POST)
		if form.is_valid():
			with transaction.atomic():
				booking = form.save(commit=False)
				booking.user = request.user
				booking.status = Booking.STATUS_BOOKED
				booking.save()
				if request.user.email:
					send_templated_email(
						subject="Booking Confirmation | SHMS",
						template_name="emails/notification.html",
						context={
							"title": "Booking confirmed",
							"message": f"Your booking for Room {booking.room.room_number} and Bed {booking.bed.bed_number} has been confirmed.",
							"action_url": request.build_absolute_uri(reverse("bookings:qr", args=[booking.id])),
						},
						recipients=[request.user.email],
					)
				create_notification(
					recipient=request.user,
					title="Booking confirmed",
					message=f"Your booking for Room {booking.room.room_number} / Bed {booking.bed.bed_number} is confirmed.",
					related_url=booking.get_absolute_url() if hasattr(booking, "get_absolute_url") else "",
				)

			messages.success(request, "Booking created successfully. QR code generated.")
			return redirect("bookings:qr", booking_id=booking.id)
	else:
		form = BookingForm()

	return render(request, "bookings/book_room.html", {"form": form, "rooms": rooms})


@login_required
def booking_history_view(request):
	if request.user.role != User.ROLE_ADMIN:
		return HttpResponseForbidden("Access denied")

	bookings = Booking.objects.select_related("user", "room", "bed", "room__hostel")
	is_admin = True
	filter_type = request.GET.get("filter") or ""
	selected_date = request.GET.get("date") or ""
	room_query = request.GET.get("room") or ""
	student_query = request.GET.get("student") or ""
	status_query = request.GET.get("status") or ""
	start_date = parse_date(request.GET.get("start_date") or "")
	end_date = parse_date(request.GET.get("end_date") or "")

	if filter_type == "today":
		bookings = bookings.filter(start_date=date.today())
	elif filter_type == "tomorrow":
		bookings = bookings.filter(start_date=date.today() + timedelta(days=1))
	elif filter_type == "custom" and selected_date:
		bookings = bookings.filter(start_date=selected_date)

	if room_query:
		bookings = bookings.filter(room__room_number__icontains=room_query)
	if student_query:
		bookings = bookings.filter(user__username__icontains=student_query)
	if status_query:
		bookings = bookings.filter(status=status_query)
	if start_date:
		bookings = bookings.filter(created_at__date__gte=start_date)
	if end_date:
		bookings = bookings.filter(created_at__date__lte=end_date)

	paginator = Paginator(bookings.order_by("-created_at"), 10)
	page_number = request.GET.get("page")
	bookings_page = paginator.get_page(page_number)

	context = {
		"bookings": bookings_page,
		"active_filter": filter_type,
		"selected_date": selected_date,
		"room_query": room_query,
		"student_query": student_query,
		"status_query": status_query,
		"start_date": start_date.isoformat() if start_date else "",
		"end_date": end_date.isoformat() if end_date else "",
		"is_admin": is_admin,
	}
	return render(request, "bookings/history.html", context)


@login_required
def booking_qr_view(request, booking_id):
	booking = get_object_or_404(Booking.objects.select_related("user", "room", "bed", "room__hostel"), pk=booking_id)

	if booking.user != request.user and request.user.role != User.ROLE_ADMIN and not request.user.is_staff:
		messages.error(request, "You are not allowed to view this booking QR.")
		return redirect("bookings:history")

	return render(request, "bookings/booking_qr.html", {"booking": booking, "today": date.today()})


@login_required
def check_in_view(request, booking_id):
	booking = get_object_or_404(Booking.objects.select_related("user", "room", "bed", "room__hostel"), pk=booking_id)

	if booking.user != request.user and request.user.role != User.ROLE_ADMIN and not request.user.is_staff:
		messages.error(request, "You are not allowed to check in for this booking.")
		return redirect("bookings:history")

	if request.method == "POST":
		today = date.today()
		if today < booking.start_date or today > booking.end_date:
			messages.warning(request, "Check-in is only available during the booking period.")
		elif booking.status != Booking.STATUS_BOOKED:
			messages.warning(request, "Only active bookings can be checked in.")
		elif booking.check_in_at:
			messages.info(request, "Check-in is already marked for this booking.")
		else:
			booking.check_in_at = timezone.now()
			booking.save(update_fields=["check_in_at"])
			messages.success(request, "Check-in marked successfully.")
		return redirect("bookings:check_in", booking_id=booking.id)

	return render(request, "bookings/check_in.html", {"booking": booking})


@login_required
def cancel_booking_view(request, booking_id):
	booking = get_object_or_404(Booking, pk=booking_id)

	if booking.user != request.user and request.user.role != User.ROLE_ADMIN and not request.user.is_staff:
		messages.error(request, "You are not allowed to cancel this booking.")
		return redirect("bookings:history")

	if booking.status == Booking.STATUS_BOOKED:
		booking.status = Booking.STATUS_CANCELLED
		booking.save(update_fields=["status"])
		messages.info(request, "Booking cancelled.")
	else:
		messages.warning(request, "Only active bookings can be cancelled.")

	return redirect("bookings:history")


@login_required
def export_bookings_excel(request):
	if not _is_admin_user(request.user):
		return HttpResponseForbidden("Access denied")

	bookings = _get_latest_unique_bookings_for_export()

	workbook = Workbook()
	worksheet = workbook.active
	worksheet.title = "Bookings"

	headers = ["Student Name", "Booked On", "Room Number", "Bed Number", "From Date", "To Date", "Status"]
	worksheet.append(headers)
	for cell in worksheet[1]:
		cell.font = Font(bold=True)

	for booking in bookings:
		worksheet.append(
			[
				booking.user.get_full_name() or booking.user.username,
				booking.created_at.strftime("%d-%m-%Y") if booking.created_at else "",
				booking.room.room_number,
				_get_export_bed_label(booking),
				booking.start_date.strftime("%d-%m-%Y"),
				booking.end_date.strftime("%d-%m-%Y"),
				booking.get_status_display(),
			]
		)

	# Auto-fit columns to keep all content fully visible.
	for col_idx, column_cells in enumerate(worksheet.columns, start=1):
		max_length = 0
		for cell in column_cells:
			value = "" if cell.value is None else str(cell.value)
			if len(value) > max_length:
				max_length = len(value)
		worksheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

	response = HttpResponse(
		content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
	)
	response["Content-Disposition"] = 'attachment; filename="bookings.xlsx"'
	workbook.save(response)
	return response
